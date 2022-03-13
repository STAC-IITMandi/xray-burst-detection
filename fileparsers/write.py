"""
All functions to write data of the detected bursts to various output formats are defined here.
"""

from astropy.io import fits as FITSio, ascii as ASCIIio
from astropy.table import Table
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numpy as np

import io
import logging

logger = logging.getLogger('fileparsers.write')

__all__ = ['fits','excel','text']


def fits(jsondata, fp, filename='', filedate=''):
    t = Table(rows=jsondata)
    u0 = FITSio.PrimaryHDU()
    u1 = FITSio.BinTableHDU(t, name='BURSTS')
    u0.header['filename'] = filename
    u0.header['uplodate'] = str(filedate)
    u0.header['software'] = 'X-Ray burst detector'
    u0.header['softwURL'] = 'https://github.com/STAC-IITMandi/xray-burst-detection'
    hdul = FITSio.HDUList([u0,u1])
    hdul.writeto(fp, overwrite=True)


def excel(jsondata, fp, filename='', filedate=''):
    wb = Workbook()
    s1 = wb.active
    s1.title = "BURSTS"
    cols = {}
    for i, colname in enumerate(jsondata[0], start=1):
        s1.cell(row=1, column=i, value=colname)
        cols[colname] = i
    for j, record in enumerate(jsondata, start=2) :
        for colname, val in record.items():
            s1.cell(row=j, column=cols[colname], value=val)
    s0 = wb.create_sheet('Metadata')
    s0['A1'], s0['B1']= 'filename', filename
    s0['A2'], s0['B2']= 'upload_date', str(filedate)
    s0['A3'], s0['B3']= 'software', 'X-Ray burst detector'
    s0['A4'], s0['B4']= 'software_URL', 'https://github.com/STAC-IITMandi/xray-burst-detection'
    wb.save(fp)


def text(jsondata, fp, filename='', filedate=''):
    t = Table(rows=jsondata)
    t.meta['filename'] = filename
    t.meta['upload_date'] = str(filedate)
    t.meta['software'] = 'X-Ray burst detector'
    t.meta['software_URL'] = 'https://github.com/STAC-IITMandi/xray-burst-detection'
    t.write(fp, format='ascii.csv', overwrite=True)



def timeseries(fp, ts, fs, er):
    t = Table([ts, fs], names=['x','y'])
    if er is not None :
        t.add_column(er, name="(_(ERRORVALS)_)")
        tempfile = io.StringIO()
        t.write(tempfile, format='ascii.csv', delimiter=',')
        clean = tempfile.getvalue().replace(",(_(ERRORVALS)_)", "")
        with open(fp, 'w') as outfile :
            outfile.write(clean)
    else :
        with open(fp, 'w') as outfile :
            t.write(outfile, format='ascii.csv', delimiter=',')
